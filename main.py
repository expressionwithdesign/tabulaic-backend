import os, re, time, uuid, tempfile
import pandas as pd
import anthropic
import stripe
 
from fastapi import FastAPI, File, UploadFile, HTTPException, Header, Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
 
app = FastAPI(title="Tabulaic API", version="3.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=False, allow_methods=["*"], allow_headers=["*"])
 
FREE_ROW_LIMIT        = 1
BATCH_SIZE            = 30
ANTHROPIC_API_KEY     = os.environ.get("ANTHROPIC_API_KEY")
STRIPE_SECRET_KEY     = os.environ.get("STRIPE_SECRET_KEY")
STRIPE_WEBHOOK_SECRET = os.environ.get("STRIPE_WEBHOOK_SECRET")
stripe.api_key = STRIPE_SECRET_KEY
 
TIER_LIMITS = {
    "price_1Tnlr9RoZgDHWVTnMoItFrlt": 100,
    "price_1TnlsGRoZgDHWVTnYQBwcujz": 1000,
    "price_1TnltNRoZgDHWVTnPtrSfee2": 999999,
}
TOKEN_STORE: dict[str, int] = {}
 
SYSTEM_PROMPT = """You are an expert Amazon Catalog Optimization Agent. Your sole task is to rewrite existing, long Amazon product titles to comply with the strict July 27, 2026 Title Update guidelines while defending organic keyword indexing.
 
STRICT CONSTRAINTS & SEQUENCING:
 
1. OPTIMIZED TITLE (Max 75 chars):
   - Mandatory Structure: [Brand Name] + [Core Product Type/Keywords] + [Dimensions/Specs] + [Pack Count].
   - CRITICAL BRAND RULE: The Brand Name MUST be the very first word in the title. If a brand name is already present inside the raw title, move it to the absolute front. If NO brand name is found anywhere in the provided text, automatically default and use "FoldCard" as the brand name prefix.
   - Aggressively remove filler: Eliminate use-case stuffing, promotional fluff, marketing adjectives (e.g., "premium", "superb"), and redundant restatements.
   - Style: Capitalize first letters of major words. Use numerals ("2") instead of words ("two"). No repeated words. No banned characters (!, $, ?, _, {, }, ^).
 
2. ITEM HIGHLIGHTS (Max 125 chars):
   - Recapture high-value search keywords dropped from the title (e.g., paper weights like 80lb/32lb, coating details like C2S, and end-use applications like brochures, flyers, crafting).
   - Format as a dense, natural, or semi-structured phrase focusing on key specs, materials, compatibility, or what's included. Do not let it overflow past 125 characters.
 
PROCESSING METHOD:
- If a spreadsheet or multi-row text list is provided, process every row sequentially and return the output in the exact same order as the input.
 
OUTPUT FORMAT:
Return exactly this clean layout for every SKU with no extra conversational preamble or generic text:
 
[ASIN / SKU]
- **New Title (X chars):** [Your optimized title]
- **Item Highlights (Y chars):** [Your item highlights text]
---"""
 
def parse_upload(file_bytes, filename):
    ext = filename.lower().rsplit(".", 1)[-1]
    with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
        tmp.write(file_bytes); tmp_path = tmp.name
    if ext in ("tsv","txt"):    df = pd.read_csv(tmp_path, sep="\t")
    elif ext == "csv":          df = pd.read_csv(tmp_path)
    elif ext in ("xlsx","xls"): df = pd.read_excel(tmp_path)
    else: raise ValueError(f"Unsupported file type: {ext}")
    os.unlink(tmp_path)
    df.columns = [c.strip() for c in df.columns]
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if cl in ("asin","asin1") and "asin" not in col_map: col_map["asin"] = col
        if cl in ("title","item-name","item_name","product title","product_title") and "title" not in col_map: col_map["title"] = col
    if "title" not in col_map: raise ValueError("Could not find a title column. Expected: 'Title', 'item-name', or 'product title'.")
    df = df.rename(columns={v: k for k, v in col_map.items()})
    df = df.dropna(subset=["title"]).reset_index(drop=True)
    if "asin" not in df.columns: df["asin"] = [f"ROW_{i+1}" for i in range(len(df))]
    return df[["asin","title"]]
 
def parse_claude_response(text):
    results = []
    title_re     = re.compile(r"\*\*New Title \((\d+) chars?\):\*\*\s*(.+)", re.IGNORECASE)
    highlight_re = re.compile(r"\*\*Item Highlights \((\d+) chars?\):\*\*\s*(.+)", re.IGNORECASE)
    for block in [b.strip() for b in text.split("---") if b.strip()]:
        lines = block.strip().splitlines()
        if not lines: continue
        asin = lines[0].strip(); new_title = highlights = ""
        for line in lines[1:]:
            tm = title_re.search(line); hm = highlight_re.search(line)
            if tm: new_title = tm.group(2).strip()
            if hm: highlights = hm.group(2).strip()
        if new_title:
            new_title  = new_title[:75]
            highlights = highlights[:125]
            results.append({
                "asin": asin, "new_title": new_title, "title_chars": len(new_title),
                "highlights": highlights, "highlight_chars": len(highlights)
            })
    return results
 
def process_with_claude(df):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    all_results = []
    for i in range(0, len(df), BATCH_SIZE):
        batch  = df.iloc[i:i+BATCH_SIZE].to_dict("records")
        prompt = "\n\n".join([f"ASIN: {r['asin']}\nTitle: {r['title']}" for r in batch])
        response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=4096,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": prompt}]
        )
        all_results.extend(parse_claude_response(response.content[0].text))
        print(f"  Processed rows {i+1}–{min(i+BATCH_SIZE, len(df))} of {len(df)}")
        if i + BATCH_SIZE < len(df): time.sleep(0.5)
    return all_results
 
def build_excel(results, output_path):
    wb = Workbook(); ws = wb.active; ws.title = "Optimized Titles"
    thin   = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill  = PatternFill("solid", start_color="1F4E79")
    hfont  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    altf   = PatternFill("solid", start_color="EBF3FB")
    dfont  = Font(name="Arial", size=10)
    green  = PatternFill("solid", start_color="C6EFCE")
    red    = PatternFill("solid", start_color="FFC7CE")
    wrap   = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top")
    for col, h in enumerate(["ASIN","New Title","Title Chars","Item Highlights","Highlights Chars"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    ws.row_dimensions[1].height = 30
    for ri, r in enumerate(results, 2):
        fill = altf if ri % 2 == 0 else None
        for col, val in enumerate([r["asin"],r["new_title"],r["title_chars"],r["highlights"],r["highlight_chars"]], 1):
            c = ws.cell(row=ri, column=col, value=val); c.font = dfont; c.border = border
            if fill: c.fill = fill
            c.alignment = center if col in (1,3,5) else wrap
        ws.cell(row=ri, column=3).fill = green if r["title_chars"] <= 75 else red
        ws.cell(row=ri, column=5).fill = green if r["highlight_chars"] <= 125 else red
    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 13; ws.column_dimensions["D"].width = 68
    ws.column_dimensions["E"].width = 17; ws.freeze_panes = "A2"
    wb.save(output_path)
 
@app.get("/")
def root(): return {"status": "Tabulaic API running"}
 
@app.get("/health")
def health(): return {"status": "ok"}
 
@app.post("/create-checkout-session")
async def create_checkout_session(request: Request):
    body = await request.json()
    price_id = body.get("price_id")
    if price_id not in TIER_LIMITS: raise HTTPException(status_code=400, detail="Invalid price ID.")
    try:
        session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            line_items=[{"price": price_id, "quantity": 1}],
            mode="payment",
            success_url="https://tabulaic.com/success?token={CHECKOUT_SESSION_ID}",
            cancel_url="https://tabulaic.com/#pricing",
            metadata={"price_id": price_id},
        )
        return JSONResponse({"checkout_url": session.url})
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))
 
@app.post("/webhook")
async def stripe_webhook(request: Request):
    payload    = await request.body()
    sig_header = request.headers.get("stripe-signature")
    try: event = stripe.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
    except stripe.error.SignatureVerificationError: raise HTTPException(status_code=400, detail="Invalid signature.")
    if event["type"] == "checkout.session.completed":
        session  = event["data"]["object"]
        price_id = session["metadata"].get("price_id")
        token    = session["id"]
        TOKEN_STORE[token] = TIER_LIMITS.get(price_id, FREE_ROW_LIMIT)
        print(f"Token issued: {token} → {TOKEN_STORE[token]} rows")
    return JSONResponse({"status": "ok"})
 
@app.post("/process")
async def process_file(file: UploadFile = File(...), x_token: str = Header(default=None)):
    if not ANTHROPIC_API_KEY: raise HTTPException(status_code=500, detail="API key not configured.")
    row_limit  = TOKEN_STORE.pop(x_token, FREE_ROW_LIMIT) if x_token and x_token in TOKEN_STORE else FREE_ROW_LIMIT
    file_bytes = await file.read()
    if not file_bytes: raise HTTPException(status_code=400, detail="Empty file.")
    try: df = parse_upload(file_bytes, file.filename)
    except ValueError as e: raise HTTPException(status_code=400, detail=str(e))
    total_rows    = len(df)
    df_to_process = df.head(row_limit)
    try: results = process_with_claude(df_to_process)
    except Exception as e: raise HTTPException(status_code=500, detail=f"Claude error: {str(e)}")
    if not results: raise HTTPException(status_code=500, detail="No results returned.")
    output_path = f"/tmp/tabulaic_{uuid.uuid4().hex[:8]}.xlsx"
    build_excel(results, output_path)
    return FileResponse(
        path=output_path, filename="tabulaic_optimized.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"X-Total-Rows": str(total_rows), "X-Processed-Rows": str(len(df_to_process))}
    )
 
@app.post("/preview")
async def preview_file(file: UploadFile = File(...)):
    file_bytes = await file.read()
    try: df = parse_upload(file_bytes, file.filename)
    except ValueError as e: raise HTTPException(status_code=400, detail=str(e))
    sample = df.head(1).to_dict("records")
    return JSONResponse({"total_rows": len(df), "sample_asin": sample[0]["asin"] if sample else "", "sample_title": sample[0]["title"] if sample else ""})
